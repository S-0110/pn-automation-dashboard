#!/usr/bin/env python
# coding: utf-8

import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

# =====================================================
# FILE PATHS
# =====================================================

# =====================================================
# FILE PICKER + OUTPUT PATHS
# =====================================================

from pathlib import Path
import sys

# Streamlit passes the uploaded file path
if len(sys.argv) < 2:
    print("Usage: pn.py <input_file.xlsx>", file=sys.stderr)
    sys.exit(1)

input_file = sys.argv[1]

input_path = Path(input_file)

excel_output = (
    input_path.parent /
    f"{input_path.stem}_Analysis.xlsx"
)

csv_output = (
    input_path.parent /
    f"{input_path.stem}_Analysis.csv"
)


# =====================================================
# LOAD RAW DATA
# =====================================================

try:
    raw_df = pd.read_excel(input_file)
except Exception as e:
    print(f"Failed to read input file: {e}", file=sys.stderr)
    sys.exit(1)

# =====================================================
# REQUIRED COLUMNS
# =====================================================

required_columns = [
    'Campaign Name',
    'Campaign ID',
    'Channel',
    'Variant',
    'Title',
    'Message',
    'Start Date',
    'Start Time',
    'Device',
    'Who query',
    'Conversion Event',
    'Run Date',
    'Total Sent(users)',
    'Total Viewed(users)',
    'Total Clicked(users)',
    'Total Delivered(users)',
    'Click through conversions',
    'Click through conversions %',
    'Total control group count',
    'Total control group conversions',
    'Total control group conversions %',
    'Total control group revenue',
    'Influenced Conversions',
    'Influenced Conversions %',
    'Influenced Revenue'
]

# =====================================================
# CHECK MISSING COLUMNS
# =====================================================

missing_cols = [
    col for col in required_columns
    if col not in raw_df.columns
]

if missing_cols:
    raise ValueError(
        f"Missing columns in input file:\n{missing_cols}"
    )

# =====================================================
# CREATE ANALYSIS DATAFRAME
# =====================================================

analysis_df = raw_df[required_columns].copy()

# =====================================================
# NUMERIC COLUMNS
# =====================================================

numeric_cols = [
    'Total Sent(users)',
    'Total Viewed(users)',
    'Total Clicked(users)',
    'Total Delivered(users)',
    'Click through conversions',
    'Total control group count',
    'Total control group conversions',
    'Total control group revenue',
    'Influenced Conversions',
    'Influenced Revenue'
]

for col in numeric_cols:
    analysis_df[col] = pd.to_numeric(
        analysis_df[col],
        errors='coerce'
    ).fillna(0)

# =====================================================
# COMBINE ANDROID + IOS
# GROUP BY CAMPAIGN ID + VARIANT
# =====================================================

analysis_df = (
    analysis_df
    .groupby(
        ['Campaign ID', 'Variant'],
        dropna=False,
        as_index=False
    )
    .agg({
        'Campaign Name': 'first',
        'Channel': 'first',
        'Title': 'first',
        'Message': 'first',
        'Start Date': 'first',
        'Start Time': 'first',
        'Who query': 'first',
        'Conversion Event': 'first',
        'Run Date': 'first',

        'Device': lambda x: ', '.join(
            sorted(
                set(
                    str(v).strip()
                    for v in x
                    if pd.notna(v)
                )
            )
        ),

        'Total Sent(users)': 'sum',
        'Total Viewed(users)': 'sum',
        'Total Clicked(users)': 'sum',
        'Total Delivered(users)': 'sum',
        'Click through conversions': 'sum',
        'Total control group count': 'sum',
        'Total control group conversions': 'sum',
        'Total control group revenue': 'sum',
        'Influenced Conversions': 'sum',
        'Influenced Revenue': 'sum'
    })
)

# =====================================================
# METRIC CALCULATIONS
# =====================================================

analysis_df['CTR'] = np.where(
    analysis_df['Total Viewed(users)'] > 0,
    analysis_df['Total Clicked(users)']
    / analysis_df['Total Viewed(users)'],
    0
)

analysis_df['Total Converted'] = (
    analysis_df['Click through conversions']
    + analysis_df['Influenced Conversions']
)

analysis_df['TC%'] = np.where(
    analysis_df['Total Viewed(users)'] > 0,
    (
        analysis_df['Total Converted']
        / analysis_df['Total Viewed(users)']
    ) * 100,
    0
)

analysis_df['CTR%'] = np.where(
    analysis_df['Total Viewed(users)'] > 0,
    (
        analysis_df['Total Clicked(users)']
        / analysis_df['Total Viewed(users)']
    ) * 100,
    0
)

analysis_df['Total CG Conv %'] = np.where(
    analysis_df['Total control group count'] > 0,
    (
        analysis_df['Total control group conversions']
        / analysis_df['Total control group count']
    ) * 100,
    0
)

analysis_df['Sent Rate'] = np.where(
    analysis_df['Total Sent(users)'] > 0,
    analysis_df['Total Converted']
    / analysis_df['Total Sent(users)'],
    0
)

analysis_df['Control Rate'] = np.where(
    analysis_df['Total control group count'] > 0,
    analysis_df['Total control group conversions']
    / analysis_df['Total control group count'],
    0
)

analysis_df['Campaign Conversion Rate (CR)'] = np.where(
    analysis_df['Total Viewed(users)'] > 0,
    analysis_df['Total Converted']
    / analysis_df['Total Viewed(users)'],
    0
)

analysis_df['Relative Lift'] = np.where(
    analysis_df['Control Rate'] > 0,
    (
        (
            analysis_df['Campaign Conversion Rate (CR)']
            - analysis_df['Control Rate']
        )
        / analysis_df['Control Rate']
    ) * 100,
    0
)

analysis_df['Incremental Conversions'] = (
    (
        analysis_df['Campaign Conversion Rate (CR)']
        - analysis_df['Control Rate']
    )
    * analysis_df['Total Viewed(users)']
)

analysis_df['Click through conversions %'] = np.where(
    analysis_df['Total Viewed(users)'] > 0,
    analysis_df['Click through conversions']
    / analysis_df['Total Viewed(users)'] * 100,
    0
)

analysis_df['Total control group conversions %'] = np.where(
    analysis_df['Total control group count'] > 0,
    analysis_df['Total control group conversions']
    / analysis_df['Total control group count'] * 100,
    0
)

analysis_df['Influenced Conversions %'] = np.where(
    analysis_df['Total Viewed(users)'] > 0,
    analysis_df['Influenced Conversions']
    / analysis_df['Total Viewed(users)'] * 100,
    0
)

# =====================================================
# CATEGORY
# =====================================================

analysis_df['Category'] = (
    analysis_df['Campaign Name']
    .astype(str)
    .str.split('_')
    .str[0]
)

# =====================================================
# SEGMENT
# =====================================================

analysis_df['Segment'] = (
    analysis_df['Campaign Name']
    .astype(str)
    .str.split('_')
    .str[-1]
)


# =====================================================
# EXPORT ANALYSIS TO CSV
# =====================================================

analysis_df.to_csv(
    csv_output,
    index=False,
    encoding="utf-8-sig"
)
# =====================================================
# EXPORT TO EXCEL
# =====================================================

with pd.ExcelWriter(
    excel_output,
    engine='openpyxl'
) as writer:
    
    analysis_df.to_excel(
        writer,
        sheet_name='Analysis',
        index=False
    )

    categories = (
        analysis_df['Category']
        .dropna()
        .unique()
    )

    used_sheet_names = set()

    for category in categories:

        cat_df = analysis_df[
            analysis_df['Category'] == category
        ]

        sheet_name = str(category)

        sheet_name = re.sub(
            r'[:\\/*?\[\]]',
            '_',
            sheet_name
        ).strip()

        if not sheet_name:
            sheet_name = "Unknown"

        sheet_name = sheet_name[:31]

        original_name = sheet_name
        counter = 1

        while sheet_name in used_sheet_names:
            suffix = f"_{counter}"
            sheet_name = (
                original_name[:31-len(suffix)]
                + suffix
            )
            counter += 1

        used_sheet_names.add(sheet_name)

        cat_df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )

    # -----------------------------------------
    # CATEGORY OVERVIEW SUMMARY
    # -----------------------------------------

    summary_sheet = "Summary"

    category_summary = (
        analysis_df
        .groupby('Category')
        .agg(
            Campaigns=('Campaign ID', 'nunique'),
            Total_Sent=('Total Sent(users)', 'sum'),
            Total_Viewed=('Total Viewed(users)', 'sum'),
            Total_Clicked=('Total Clicked(users)', 'sum'),
            Total_Converted=('Total Converted', 'sum'),
            Influenced_Revenue=('Influenced Revenue', 'sum'),
            Incremental_Conversions=('Incremental Conversions', 'sum')
        )
        .reset_index()
    )

    category_summary['View Rate %'] = np.where(
        category_summary['Total_Sent'] > 0,
        category_summary['Total_Viewed']
        / category_summary['Total_Sent'] * 100,
        0
    )

    category_summary['CTR %'] = np.where(
        category_summary['Total_Viewed'] > 0,
        category_summary['Total_Clicked']
        / category_summary['Total_Viewed'] * 100,
        0
    )

    category_summary['Conversion Rate %'] = np.where(
        category_summary['Total_Viewed'] > 0,
        category_summary['Total_Converted']
        / category_summary['Total_Viewed'] * 100,
        0
    )

    category_summary = category_summary.sort_values(
        by='Total_Converted',
        ascending=False
    )

    category_summary = category_summary[
        [
            'Category',
            'Campaigns',
            'Total_Sent',
            'Total_Viewed',
            'View Rate %',
            'Total_Clicked',
            'CTR %',
            'Total_Converted',
            'Conversion Rate %',
            'Influenced_Revenue',
            'Incremental_Conversions'
        ]
    ]

    category_summary.to_excel(
        writer,
        sheet_name=summary_sheet,
        index=False
    )

# =====================================================
# FORMAT EXCEL
# =====================================================

wb = load_workbook(excel_output)

# Move Summary sheet to first position
if "Summary" in wb.sheetnames:
    summary_sheet = wb["Summary"]
    wb.move_sheet(summary_sheet, offset=-wb.index(summary_sheet))

for sheet in wb.sheetnames:

    ws = wb[sheet]
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for column_cells in ws.columns:

        max_length = max(
            len(str(cell.value))
            if cell.value is not None
            else 0
            for cell in column_cells
        )

        adjusted_width = min(
            max_length + 5,
            50
        )

        ws.column_dimensions[
            column_cells[0].column_letter
        ].width = adjusted_width

wb.save(excel_output)
